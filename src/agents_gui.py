# -*- coding:utf-8 -*-
import json
import os
import sys
import dateutil
import numpy as np
import openpyxl
import pandas
from dateutil import parser
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from agents_bbdd import AlumnesBbdd, RegistresBbdd, CategoriesBbdd, DatesBbdd, Iniciador, Liquidador
from formats import DataGuiComm, Registresguicomm, Alumne_comm, Registres_gui_nou, \
    Registres_bbdd_nou, RegistresBbddComm, AlumneNou, DataNova, CategoriaNova


class Comprovador:
    """Comprova la presencia de taules a la base de dades"""

    def __init__(self, modegui):
        super(Comprovador, self).__init__()
        resultat_iniciador = Iniciador(modegui)
        estat_alumnes = resultat_iniciador.presencia_taula_alumne
        estat_registres = resultat_iniciador.presencia_taula_registres
        estat_categories = resultat_iniciador.presencia_taula_categories
        estat_dates = resultat_iniciador.presencia_taula_dates
        estat_global = estat_alumnes and estat_registres and estat_categories and estat_dates
        if estat_global is False:
            Iniciador(1).crea_taules()


class Destructor:
    """Classe per a eliminar totes les taules de la base de dades i la base de dades en si.
    :param modegui: int 1 per a l'execucio normal, 0 per a l'execucio en mode test
    """

    def __init__(self, modegui):
        super(Destructor).__init__()
        self.alumnes = AlumnesBbdd(modegui)
        self.registres = RegistresBbdd(modegui)
        self.categories = CategoriesBbdd(modegui)
        self.dates = DatesBbdd(modegui)
        self.liquidador = Liquidador(modegui)

    def destruir(self):
        """Elimina les taules de la base de dades i, despres, la base de dades en si."""
        self.registres.destruir_taula()
        self.alumnes.destruir_taula()
        self.categories.destruir_taula()
        self.dates.destruir_taula()
        self.liquidador.eliminar_basededades()


class Comptable:
    """S'ocupa de gestionar amb la gui les entrades a la taula de registres"""

    def __init__(self, modegui):
        """Ordenar les taules de la base de dades perquè siguin més fàcils de llegir:"""
        self.info_categories = None
        self.info_registres = None
        self.info_alumnes = None
        self.alumnes = AlumnesBbdd(modegui)
        self.registrador = RegistresBbdd(modegui)
        self.categoritzador = CategoriesBbdd(modegui)
        self.info_categories = self.categoritzador.lectura_categories()
        self.info_alumnes = self.alumnes.llegir_alumnes()
        self.info_registres = self.registrador.lectura_registres()
        self.registres = self.obtenir_registres()
        self.noms_alumnes = self.obtenir_noms()
        self.categories = self.obtenir_categories()

    def obtenir_registres(self):
        """Retorna una llista de registres convenients per a la presentació: list:
        """
        # Condicio de seguretat per si no hi ha registres:
        if self.info_registres and self.info_alumnes and self.info_categories:
            registres_entrada = self.info_registres
            missatge_registres = []
            # Substituim l'id de l'alumne de la llista de registres pel seu nom:
            for registre in registres_entrada:
                registre_proces = [registre.id]
                # Obtenim el nom de l'alumne:
                registre_proces.append(self.obtenir_registres_alumne(id_alumne=registre.alumne))
                # Substituim l'id de la categoria de la llista de registres pel seu nom:
                registre_proces.append(self.obtenir_categoria_registres(id_categoria=registre.categoria))
                registre_proces.append(registre.data)
                registre_proces.append(registre.descripcio)
                registre_tractat = Registresguicomm(registre_proces[0], registre_proces[1], registre_proces[2],
                                                    registre_proces[3], registre_proces[4])
                missatge_registres.append(registre_tractat)

            return missatge_registres
        return False

    def obtenir_registres_alumne(self, id_alumne):
        """Retorna el Alumnecomm corresponent a l'id alumne proporcionat"""
        resposta = None
        for alumne in self.info_alumnes:
            if id_alumne == alumne.id:
                resposta = alumne
        return resposta

    def obtenir_categoria_registres(self, id_categoria):
        """Retorna la Categoriacomm corresponent a l'id alumne proporcionat"""
        for categoria in self.info_categories:
            if id_categoria == categoria.id:
                return categoria

    def obtenir_noms(self):
        """Retorna una llista d'alumnes convenients per a la presentació: list:"""
        self.info_alumnes = self.alumnes.llegir_alumnes()
        # Condicio de seguretat per si no hi ha alumnes:
        if self.info_alumnes:
            noms = [alumne.nom for alumne in self.info_alumnes]
            return noms
        return False

    def obtenir_categories(self):
        """Retorna una llista de categories convenients per a la presentació: list:"""
        self.info_categories = self.categoritzador.lectura_categories()
        # Condicio de seguretat per si no hi ha categories:
        if self.info_categories:
            categories = [categoria.nom for categoria in self.info_categories]
            return categories

        return False

    def actualitzar_registres(self, registre_input):
        """Processa el missatge per actualitzar-lo a la base de dades"""
        if not isinstance(registre_input, list):
            raise TypeError("El registre_input no és del tipus correcte.")
        missatge_actualitzar_registre = []
        for element in registre_input:
            if not isinstance(element, Registresguicomm):
                raise TypeError("Registre no segueix el format establert")
            registre_enviar = RegistresBbddComm(element.id, element.alumne.id, element.categoria.id, element.data,
                                                element.descripcio)
            missatge_actualitzar_registre.append(registre_enviar)
        self.registrador.actualitzar_registre(missatge_actualitzar_registre)

    def eliminar_registre(self, registre_input):
        """Processa el missatge per eliminar-lo a la base de dades"""
        if not isinstance(registre_input, list):
            raise TypeError("El registre_input no és del tipus correcte.")

        missatge_eliminar_registre = []
        for element in registre_input:
            if not isinstance(element, Registresguicomm):
                raise TypeError("Registre no segueix el format establert")
            element_processat = list
            element_processat.append(element.id, element.alumne.id, element.categoria.id, element.data,
                                     element.descripcio)
            registre_enviar = RegistresBbddComm(dada for dada in element_processat)
            missatge_eliminar_registre.append(registre_enviar)
        self.registrador.eliminar_registre(missatge_eliminar_registre)

    def crear_registre(self, registre_input):
        """Crea un registre a la base de dades:"""
        if not isinstance(registre_input, list):
            return False

        missatge_registre_bbddd = []
        # Processem els registres nous:
        for info in registre_input:
            if not isinstance(info, Registres_gui_nou):
                return False

            # Convertim la data a un format que pugui ser llegit per la base de dades:
            registrenou = Registres_bbdd_nou(info.alumne.id, info.categoria.id, info.data, info.descripcio)
            if not isinstance(registrenou.alumne, int) or not isinstance(registrenou.categoria, int):
                return False

            missatge_registre_bbddd.append(registrenou)
        self.registrador.crear_registre(missatge_registre_bbddd)
        return True

    def refrescar_registres(self):
        """Tronar a consultar els registres a peticio de la gui, per reflectir els canvis en la base de dades"""
        self.info_categories = self.categoritzador.lectura_categories()
        self.info_alumnes = self.alumnes.llegir_alumnes()
        self.info_registres = self.registrador.lectura_registres()
        self.registres = self.obtenir_registres()
        self.noms_alumnes = self.obtenir_noms()
        self.categories = self.obtenir_categories()


class Calendaritzador:
    """Classe que gestiona les dates"""

    def __init__(self, modegui):
        self.datador = DatesBbdd(modegui)
        self.info_dates = self.datador.lectura_dates()
        self.dates = self.conversio_dates()
        self.maxim_id = self.datador.maxim_id_data()

    def conversio_dates(self):
        """Obte els registres de dates, els transforma a llista i converteix dates a format dia/mes/any, basades en
        classe Data """
        if self.datador.lectura_dates():
            self.info_dates = self.datador.lectura_dates()
            for data in self.info_dates:
                data.dia = dateutil.parser.parse(data.dia).strftime("%d/%m/%Y")
            return self.info_dates
        return False

    def registra_dates(self, aniversari: list):
        """Crea nous registres a la base de dades amb la data subministrada
        :type aniversari: DataGuiComm
        :args:
        list
        :returns:
        Veritat si ha pogut crear els registres, Fals si no.
        """
        if isinstance(aniversari, list):
            for element in aniversari:
                if isinstance(element, DataNova):
                    element.dia = dateutil.parser.parse(element.dia).strftime("%Y-%m-%d")
                else:
                    return False
            self.datador.crear_data(aniversari)
        else:
            return False

    def actualitza_dates(self, aniversari):
        """Actualitza la base de dades amb la data subministrada, comprovant si es una data nova o no
        :type aniversari: DataGuiComm
        """
        if isinstance(aniversari, list):
            for element in aniversari:
                if isinstance(element, DataGuiComm):
                    element.dia = dateutil.parser.parse(element.dia).strftime("%Y-%m-%d")
                else:
                    return False
            self.datador.actualitzar_data(aniversari)
        else:
            return False


class CapEstudis:
    """S'ocupa de les gestions de la taula alumnes amb la interficie grafica"""

    def __init__(self, modegui: int):
        """

        :type modegui: int
        """
        mode = modegui
        self.alumnes = AlumnesBbdd(modebbdd=mode)
        self.registres = RegistresBbdd(mode)
        self.info_alumnes = self.alumnes.llegir_alumnes()
        self.info_alumnes_registrats = self.registres.lectura_alumnes_registrats()
        self.alumnat = self.obtenir_alumnes()
        self.alumnat_registres = self.obtenir_alumnes_registrats()

    def obtenir_alumnes(self):
        """Retorna una llista d'alumnes amb el format Alumne: list:"""
        # Condicio de seguretat per si no hi ha alumnes:
        alumnes_formatats = []
        if self.info_alumnes:
            for alumne in self.info_alumnes:
                alumnes_formatats.append(Alumne_comm(alumne.id, alumne.nom))
            return alumnes_formatats

        return False

    def obtenir_alumnes_registrats(self):
        """Retorna una llista d'alumnes amb el format Alumne: list:"""
        # Condicio de seguretat per si no hi ha alumnes:
        alumnes_formatats = []
        if self.info_alumnes and self.info_alumnes_registrats:
            alumnes_registrats = self.info_alumnes_registrats
            for persona in alumnes_registrats:
                alumnes_formatats.append(Alumne_comm(persona.id, persona.nom))
            return alumnes_formatats
        return False

    def eliminar_alumnes(self, llista_eliminar):
        """Elimina un alumne de la base de dades"""
        if isinstance(llista_eliminar, list):
            self.alumnes.eliminar_alumne(llista_eliminar)
            self.registres.eliminar_registre_alumne(llista_eliminar)
            return True
        return False

    def actualitzar_alumnes(self, llista_actualitzar: list):
        """Actualitza els alumnea de la base de dades"""
        if isinstance(llista_actualitzar, list):
            self.alumnes.actualitzar_alumne(llista_actualitzar)
            return True
        return False

    def afegir_alumnes(self, missatge_afegir: list):
        """Afegeix un alumne a la base de dades"""
        if not isinstance(missatge_afegir, list):
            return False

        for persona in missatge_afegir:
            if not isinstance(persona, AlumneNou):
                return False
        resultat = self.alumnes.registrar_alumne(missatge_afegir)
        return resultat

    def obtenir_id_alumne(self, alumne):
        """Retorna l'id d'un alumne"""
        if isinstance(alumne, Alumne_comm):
            for alumne_bbdd in self.info_alumnes:
                if alumne_bbdd[1] == alumne.nom:
                    return alumne_bbdd[0]
        return False

    def refrescar_alumnes(self):
        """Trona a consultar la informacio dels alumnes de la base de dades per a tenir en compte els canvis."""
        self.info_alumnes = self.alumnes.llegir_alumnes()
        self.info_alumnes_registrats = self.registres.lectura_alumnes_registrats()
        self.alumnat = self.obtenir_alumnes()
        self.alumnat_registres = self.obtenir_alumnes_registrats()


class Classificador:
    """S'ocupa de les categories a la base de dades."""

    def __init__(self, modegui):
        self.categoritzador = CategoriesBbdd(modegui)
        self.info_categories = self.categoritzador.lectura_categories()
        self.categories = self.info_categories
        self.registrador = RegistresBbdd(modegui)
        self.info_registres = self.registrador.lectura_registres()
        self.categories_registrades = self.obtenir_categories_registrades()

    def obtenir_categories_registrades(self):
        """Determina quines categories tenen alguna entrada a la taula registres."""
        if self.info_categories and self.info_registres:
            classificacio = self.info_categories
            info_registres = self.info_registres
            llista_categories_amb_registre = []
            for element in classificacio:
                for item in info_registres:
                    if element.id == item.categoria and element not in llista_categories_amb_registre:
                        llista_categories_amb_registre.append(element)
            return llista_categories_amb_registre
        return False

    def refrescar_categories_registres(self):
        """Força actualitzacio de les categories que tenen alguna entrada a la taula de registres."""
        self.info_registres = self.registrador.lectura_registres()
        self.info_categories = self.categoritzador.lectura_categories()


def format_categories(ruta_arxiu: str):
    """
    Modifica el format de l'arxiu d'entrada al format predeterminat
    :argument ruta_arxiu: ruta completa fins a l'arxiu
    :returns: Arxiu a la mateixa localitzacio modificat
    """
    full_calcul = openpyxl.load_workbook(ruta_arxiu)
    fulla = full_calcul.active
    noms = NamedStyle(name="noms")
    noms.font = Font(size=11, name="Calibri")
    noms.alignment = Alignment(horizontal="left", vertical="justify", wrap_text=False)
    vora_simple = Side(border_style='thin')
    noms.border = Border(top=vora_simple, right=vora_simple, bottom=vora_simple, left=vora_simple)
    noms.width = 20
    titols = NamedStyle(name="titols")
    titols.font = Font(size=11, name="Calibri", bold=True)
    titols.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    titols.border = Border(top=vora_simple, right=vora_simple, bottom=vora_simple, left=vora_simple)
    fulla.column_dimensions['A'].width = 20

    for cell in fulla['B']:
        cell.style = noms
    for cell in fulla['A']:
        cell.style = titols
    fulla['A1'].value = "Mesos"
    fulla['A1'].style = titols
    fulla['B1'].value = "Alumnes"
    fulla['B1'].style = titols
    fulla.column_dimensions['B'].width = 20
    full_calcul.save(ruta_arxiu)


def format_alumnes(ruta_arxiu: str):
    """Dona format a l'informe d'alumnes"""
    full_calcul = openpyxl.load_workbook(ruta_arxiu)
    color_taronja = "F6B26B"
    fulla = full_calcul.active
    llista_columnes = ['B', 'C', 'D', 'E', 'F']
    # Definim els estils:
    titols_trimestres = NamedStyle(name="titols_trimestres")
    titols_trimestres.font = Font(size=12, name="Arial", bold=True)
    titols_trimestres.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
    vora_simple = Side(border_style='thin')
    titols_trimestres.border = Border(top=vora_simple, right=vora_simple, bottom=vora_simple, left=vora_simple)
    titols = NamedStyle(name="titols")
    titols.font = Font(size=12, name="Arial", bold=True)
    titols.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
    vora_simple = Side(border_style='thin')
    titols.border = Border(top=vora_simple, right=vora_simple, bottom=vora_simple, left=vora_simple)

    titols.fill = PatternFill(fill_type='solid', start_color=color_taronja, end_color=color_taronja)
    continguts = NamedStyle(name="títols")
    continguts.font = Font(size=10, name="Arial", bold=False)
    continguts.alignment = Alignment(horizontal="justify", vertical="top", wrap_text=False)
    vora_simple = Side(border_style='thin')
    continguts.border = Border(top=vora_simple, right=vora_simple, bottom=vora_simple, left=vora_simple)
    for cell in fulla['A']:
        cell.style = titols_trimestres
    for cell in fulla['1']:
        cell.style = titols
    fulla.column_dimensions['A'].width = 13
    for column in llista_columnes:
        fulla.column_dimensions[column].width = 30

    llista_valors = []
    n_index = 0
    # Fusionem les cel·les si tenen el mateix valor:
    for cell in fulla['A']:
        referencia = [cell.value, cell.row]
        llista_valors.append(referencia)
        if n_index != 0 and cell.value == llista_valors[n_index - 1][0]:
            cell.value = ""
            fulla.merge_cells(start_row=llista_valors[n_index - 1][1], start_column=cell.column, end_row=cell.row,
                              end_column=cell.column)
            fulla.cell(row=llista_valors[n_index - 1][1], column=cell.column).style = titols_trimestres
        n_index += 1
    # Donem format a la resta de les cel·les:
    for row in fulla.iter_rows(min_row=2, max_row=fulla.max_row, min_col=2, max_col=fulla.max_column):
        for cell in row:
            cell.style = continguts
    full_calcul.save(ruta_arxiu)


class CreadorInformes:
    """Classe que crea un informe i l'exporta a la destinacio indicada

    llista_alumnes: lista d'alumnes
    :arguments:llista_categories: lista de categories
    :arguments:llista_registres: lista de registres
    :arguments:destinacio: string amb la destinacio de l'informe
    """

    def __init__(self, alumnes: list, categories: list, registres: list, carpeta_destinacio: str):
        super().__init__()
        self.dates = None
        self.alumnes = alumnes
        self.categories = categories
        self.registres = registres
        self.desti = carpeta_destinacio
        self.mesos_escolars = ["Setembre", "Octubre", "Novembre", "Desembre", "Gener", "Febrer", "Març", "Abril",
                               "Maig", "Juny", "Juliol"]
        self.mesos = ['Gener', 'Febrer', 'Març', 'Abril', 'Maig', 'Juny', 'Juliol', 'Agost', 'Setembre', 'Octubre',
                      'Novembre', 'Desembre']
        self.nombre_mesos = list(range(1, 13))

    def mes_a_string(self, mes: int):
        """Retorna el nom del mes"""
        posicio = self.nombre_mesos.index(mes)
        return self.mesos[posicio]

    def data_a_trimestre(self, data: str):
        """Retorna el trimestre de la data"""
        resultat = None
        inici_segon = dateutil.parser.parse(self.dates[0].dia, dayfirst=True)
        inici_tercer = dateutil.parser.parse(self.dates[1].dia, dayfirst=True)
        data_calculada = dateutil.parser.parse(data)
        if data_calculada < inici_segon:
            resultat = 1
        elif inici_segon <= data_calculada < inici_tercer:
            resultat = 2
        elif data_calculada >= inici_tercer:
            resultat = 3
        return resultat

    def export_categories(self):
        """Crea un informe amb les categories"""

        if self.categories and self.registres and self.alumnes:
            # Creem una llista amb les categories i els registres associats:
            llista_categories_informes = []
            llista_categories_als_registres = [registre.categoria.id for registre in self.registres]
            categories_registres = set(llista_categories_als_registres)
            for categoria in self.categories:
                if categoria.id in categories_registres:
                    llista_dades = [f"{categoria.nom}"]
                    diccionari_provisional = {"mesos": [], "alumnes": []}
                    for registre in self.registres:
                        if registre.categoria.id == categoria.id:
                            mes = dateutil.parser.parse(registre.data).month
                            mes = self.mes_a_string(mes)
                            alumne = registre.alumne.nom
                            diccionari_provisional['mesos'].append(mes)
                            diccionari_provisional['alumnes'].append(alumne)
                    Df = pandas.DataFrame(diccionari_provisional)
                    llista_dades.append(Df)
                    llista_categories_informes.append(llista_dades)
            for element in llista_categories_informes:
                agrego_mesos = {'alumnes': np.unique}
                nom_categoria = element[0]
                taula_pandas = element[1]
                taula_pandas = taula_pandas.groupby(['mesos']).aggregate(agrego_mesos).reindex(self.mesos_escolars)
                taula_pandas['alumnes'].astype(str)
                taula_pandas['alumnes'] = taula_pandas['alumnes'].str.join(', ')
                ruta_exportacio = os.path.join(self.desti, f"{nom_categoria}.xlsx")
                taula_pandas.to_excel(ruta_exportacio, index=True)
                # Apliquem format als informes:
                format_categories(ruta_exportacio)
            return True

        return False

    def export_alumne(self, dates):
        """
        Crea un informe per cada alumne
        :rtype: Un arxiu en format excel a la carpeta seleccionada per l'usuari
        """
        self.dates = dates
        if self.alumnes and self.registres and self.categories:
            llista_alumnes_informes = []
            llista_alumnes_als_registres = [registre.alumne.id for registre in self.registres]
            noms_categories = [categoria.nom for categoria in self.categories]
            alumnes_registres = set(llista_alumnes_als_registres)
            for alumne in self.alumnes:
                if alumne.id in alumnes_registres:
                    # Creem una llista amb les categories i els registres associats, una per alumne:
                    llista_dades = [f"{alumne.nom}"]
                    diccionari_provisional = {'Trimestre': []}
                    # Afegim les categories al diccionari provisional:

                    for item in noms_categories:
                        diccionari_provisional[item] = []
                    for registre in self.registres:
                        # Si el registre pertany a l'alumne, l'afegim al diccionari provisional:
                        if registre.alumne.id == alumne.id:
                            trimestre = self.data_a_trimestre(registre.data)
                            data_python = dateutil.parser.parse(registre.data)
                            data_format = data_python.strftime('%d/%m/%Y')
                            text_afegir = f"{data_format} - {registre.descripcio}"
                            # Afegim la data i la descripció al diccionari provisional si coincideix amb la categoria:
                            # Si la categoria es diferent, afegim un element buit, per poder tractar-lo amb el pandas.
                            n_reg_trimestres = len(diccionari_provisional['Trimestre'])
                            n_reg_categoria = len(diccionari_provisional[registre.categoria.nom])
                            # Comprovacio de si el trimestre es diferent. Si no, afegim el text a afegir. Si es
                            # diferent, afegim el text i el trimestre.
                            if n_reg_trimestres == 0:
                                diccionari_provisional['Trimestre'].append(trimestre)
                            else:
                                if diccionari_provisional['Trimestre'][n_reg_trimestres - 1] != trimestre:
                                    for categoria in noms_categories:
                                        while len(diccionari_provisional[categoria]) < n_reg_trimestres:
                                            diccionari_provisional[categoria].append('')
                                    diccionari_provisional['Trimestre'].append(trimestre)
                                elif n_reg_trimestres <= n_reg_categoria:
                                    diccionari_provisional['Trimestre'].append(trimestre)
                            diccionari_provisional[registre.categoria.nom].append(text_afegir)

                    # A continuacio ens assegurem que totes les categories tenen el mateix nombre d'elements:
                    for categoria in noms_categories:
                        nombre_registres = len(diccionari_provisional['Trimestre'])
                        while len(diccionari_provisional[categoria]) < nombre_registres:
                            diccionari_provisional[categoria].append('')
                    # Afegim el resultat a la llista d'alumnes:
                    Df = pandas.DataFrame(diccionari_provisional).fillna('')
                    Df['Trimestre'] = Df['Trimestre'].astype("category")
                    llista_dades.append(Df)
                    llista_alumnes_informes.append(llista_dades)
            for element in llista_alumnes_informes:
                nom_alumne = element[0]
                dades = element[1]
                ruta_exportacio = os.path.join(self.desti, f"{nom_alumne}.xlsx")
                dades.to_excel(ruta_exportacio, index=False, sheet_name="Informe", header=True, merge_cells=True)
                # Apliquem format als informes:
                format_alumnes(ruta_exportacio)
            return True
        return False


class ExportadorImportador:
    def __init__(self, mode):
        seleccio = AlumnesBbdd(mode).llegir_alumnes()
        self.registres = RegistresBbdd(mode).lectura_registres()
        self.categories = CategoriesBbdd(mode).lectura_categories()
        self.noms_alumnes = [(alumne.id, alumne.nom) for alumne in seleccio]
        self.dict_json = {"dades": []}
        self.alumnes_registrats = [alumne.nom for alumne in AlumnesBbdd(1).llegir_alumnes()]
        self.categories_registrades = [categoria.nom for categoria in CategoriesBbdd(1).lectura_categories()]
        self.nous_alumnes = []
        self.noves_categories = []
        self.nous_registres = []
        self.registres_tractar = []

    def obtencio_id_alumne(self, nom: str):
        alumne_complet = None
        alumnes_format_comm = AlumnesBbdd(1).llegir_alumnes()
        for persona in alumnes_format_comm:
            if persona.nom == nom:
                alumne_complet = persona
        nombre_alumne = alumne_complet.id
        return nombre_alumne

    def obtencio_id_categoria(self, nom: str):
        categoria_completa = None
        categories_format = CategoriesBbdd(1).lectura_categories()
        for categoria in categories_format:
            if categoria.nom == nom:
                categoria_completa = categoria
        nombre_categoria = categoria_completa.id
        return nombre_categoria

    def exportacio(self, llista_alumnes: list):
        for alumne in llista_alumnes:
            estructura = {}
            registres_alumne = []
            for registre in self.registres:
                if registre.alumne == alumne.id:
                    dades_afegir = [registre.categoria, registre.data, registre.descripcio]
                    for categoria in self.categories:
                        if registre.categoria == categoria.id:
                            dades_afegir[0] = categoria.nom
                    registres_alumne.append(dades_afegir)
            estructura["nom"] = alumne.nom
            estructura["registres"] = registres_alumne
            self.dict_json["dades"].append(estructura)
            with open(f"{alumne.nom}", "w", encoding='utf-8') as file:
                json.dump(self.dict_json, file, indent=4, ensure_ascii=False)
                file.close()

    def importacio(self, arxiu: str):
        with open(f"{arxiu}", "r", encoding='utf-8') as file:
            lectura_json = json.load(file)
            file.close()
        for alumne in lectura_json["dades"]:
            for registre in alumne["registres"]:
                registre.insert(0, alumne["nom"])
                self.registres_tractar.append(registre)
        for registre in self.registres_tractar:
            if registre[0] not in self.alumnes_registrats:
                self.nous_alumnes.append(AlumneNou(registre[0]))
            if registre[1] not in self.categories_registrades:
                self.noves_categories.append(CategoriaNova(registre[1].strip()))
        # Creem nous valors d'alumne i de categories si s'escau, per a complir amb la restriccio de la base de dades:
        if len(self.nous_alumnes) > 0:
            AlumnesBbdd(1).registrar_alumne(self.nous_alumnes)
        if len(self.noves_categories) > 0:
            CategoriesBbdd(1).crear_categoria(self.noves_categories)
        for registre in self.registres_tractar:
            # Intercanviem el nom de l'alumne pel seu id a la base de dades:
            registre[0] = self.obtencio_id_alumne(registre[0])
            # Intercanviem el nom de la categoria pel seu id a la base de dades:
            registre[1] = self.obtencio_id_categoria(registre[1])
            self.nous_registres.append(Registres_bbdd_nou(registre[0], registre[1], registre[2], registre[3]))
        # RegistresBbdd(1).crear_registre(registres_tractar)
