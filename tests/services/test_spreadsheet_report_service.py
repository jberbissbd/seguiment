from openpyxl import load_workbook
from datetime import date
from types import SimpleNamespace
import pytest
import unicodedata

from tutopy.application import create_services
from tutopy.models.messaging import (
    CategoryNew, NoteNew, StudentGroupHistoryNew, StudentNew,
)
from tutopy.models.reporting import TermConfigurationNew
from tutopy.services.exceptions import EntityNotFoundError, ValidationError
from tutopy.services.spreadsheet_report_service import SpreadsheetReportService


def _report_scenario(db, tmp_path):
    services = create_services(db)
    student = db.students.create(StudentNew("Laia", "Martí", "4t B"))
    academic = db.categories.create(CategoryNew("Acadèmic"))
    family = db.categories.create(CategoryNew("Família"))
    conduct = db.categories.create(CategoryNew("Conducta"))
    services.report_configuration.set_category_order([
        academic.id, family.id, conduct.id,
    ])
    course = db.academic_courses.get_or_create("2025-2026")
    db.student_group_history.create(StudentGroupHistoryNew(
        student.id, "4t A", "2025-09-01", "2026-02-01", course.id
    ))
    db.student_group_history.create(StudentGroupHistoryNew(
        student.id, "4t B", "2026-02-01", None, course.id
    ))
    services.report_configuration.save_term_configuration(TermConfigurationNew(
        course.id, "4t A", "2026-01-08", "2026-04-07"
    ))
    services.report_configuration.save_term_configuration(TermConfigurationNew(
        course.id, "4t B", "2026-01-10", "2026-04-10"
    ))
    for category, note_date, content in (
        (family, "2025-09-15", "Entrevista inicial"),
        (academic, "2026-01-15", "Bona evolució"),
        (conduct, "2026-02-10", "Incidència resolta"),
        (academic, "2026-04-15", "Objectius assolits"),
    ):
        services.notes.create(NoteNew(student.id, category.id, note_date, 0, content))
    return services, student, (academic, family, conduct), tmp_path / "informe.xlsx"


def test_exporta_full_per_curs_amb_categories_i_grup_historic(db, tmp_path):
    services, student, _categories, destination = _report_scenario(db, tmp_path)
    path = services.spreadsheet_reports.export_student(
        student.id, destination, include_terms=True
    )
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["2025-2026"]
    sheet = workbook["2025-2026"]
    assert sheet["A1"].value == "Martí, Laia — 4t A, 4t B"
    assert sheet["A2"].value == (
        f"Data d’exportació: {date.today().strftime('%d/%m/%Y')}"
    )
    assert [sheet.cell(3, column).value for column in range(1, 6)] == [
        "Trimestre", "Grup", "Acadèmic", "Família", "Conducta"
    ]
    assert sheet["A4"].value == "1r"
    assert sheet["B4"].value == "4t A"
    assert sheet["D4"].value == "15/09/2025 - Entrevista inicial"
    assert sheet["C5"].value == "15/01/2026 - Bona evolució"
    assert sheet["B6"].value == "4t B"
    assert sheet["E6"].value == "10/02/2026 - Incidència resolta"
    assert sheet["C7"].value == "15/04/2026 - Objectius assolits"
    assert "A5:A6" in {str(area) for area in sheet.merged_cells.ranges}
    assert sheet.freeze_panes == "A4"


def test_exportacio_sense_trimestres_no_crea_la_columna(db, tmp_path):
    services, student, _categories, _destination = _report_scenario(db, tmp_path)
    path = services.spreadsheet_reports.export_student(
        student.id, tmp_path / "sense-trimestres", include_terms=False
    )
    sheet = load_workbook(path).active
    assert path.suffix == ".xlsx"
    assert [sheet.cell(3, column).value for column in range(1, 5)] == [
        "Grup", "Acadèmic", "Família", "Conducta"
    ]
    assert not any(str(area).startswith("A4:A") for area in sheet.merged_cells.ranges)


def test_crea_un_full_per_curs_en_ordre_ascendent(db, tmp_path):
    services, student, categories, destination = _report_scenario(db, tmp_path)
    services.notes.create(NoteNew(
        student.id, categories[0].id, "2024-10-01", 0, "Curs anterior"
    ))
    workbook = load_workbook(services.spreadsheet_reports.export_student(
        student.id, destination
    ))
    assert workbook.sheetnames == ["2024-2025", "2025-2026"]


def test_rebutja_alumne_inexistent_o_sense_notes(db, tmp_path):
    services = create_services(db)
    with pytest.raises(EntityNotFoundError):
        services.spreadsheet_reports.export_student(999, tmp_path / "x.xlsx")
    student = services.students.create(StudentNew("Pau", "Puig", "3r A"))
    with pytest.raises(ValidationError, match="no té notes"):
        services.spreadsheet_reports.export_student(student.id, tmp_path / "x.xlsx")


def test_saneja_unicode_controls_formules_i_longitud(db, tmp_path):
    services = create_services(db)
    decomposed_name = unicodedata.normalize("NFD", "Júlia")
    student = services.students.create(StudentNew(decomposed_name, "Nuñez 😊", "4t\x00 A"))
    category = services.categories.create(CategoryNew("=HYPERLINK(\"x\")"))
    long_content = "Evolució\x07 positiva 😊 " + ("à" * 40_000)
    services.notes.create(NoteNew(
        student.id, category.id, "2026-02-01", 0, long_content
    ))
    sheet = load_workbook(services.spreadsheet_reports.export_student(
        student.id, tmp_path / "unicode.xlsx"
    )).active
    assert sheet["A1"].value.startswith("Nuñez 😊, Júlia — 4t A")
    assert "\x00" not in sheet["A1"].value
    assert sheet["B3"].value == '=HYPERLINK("x")'
    assert sheet["B3"].data_type == "s"
    assert "\x07" not in sheet["B4"].value
    assert "😊" in sheet["B4"].value
    assert len(sheet["B4"].value) == 32_767
    assert sheet["B4"].data_type == "s"


def _note(note_id, note_date):
    return SimpleNamespace(id=note_id, date=note_date)


def _history(history_id, group_name, start_date, end_date=None):
    return SimpleNamespace(
        id=history_id, group_name=group_name,
        start_date=start_date, end_date=end_date,
    )


def test_group_for_notes_gestiona_buits_i_solapaments():
    notes = [
        _note(1, "2025-08-01"),  # abans de qualsevol tram: usa el fallback
        _note(2, "2025-09-15"),  # dins del primer tram
        _note(3, "2026-01-20"),  # en un buit entre trams: usa el fallback
        _note(4, "2026-03-01"),  # dins del segon tram
        _note(5, "2026-03-10"),  # coincidència de dates: desempat per start_date/id
    ]
    histories = [
        _history(1, "3r A", "2025-09-01", "2025-12-31"),
        _history(2, "3r B", "2026-02-01", None),
        _history(3, "3r B (repetit)", "2026-03-05", None),
    ]

    result = SpreadsheetReportService._group_for_notes(notes, histories, "Sense grup")

    assert result[1] == "Sense grup"
    assert result[2] == "3r A"
    assert result[3] == "Sense grup"
    assert result[4] == "3r B"
    assert result[5] == "3r B (repetit)"
