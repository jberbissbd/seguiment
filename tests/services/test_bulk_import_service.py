from openpyxl import load_workbook
from odf import table, text
from odf.opendocument import OpenDocumentSpreadsheet
import pytest

from tutopy.application import create_services
from tutopy.models.bulk_import import ImportAction, ImportDecision
from tutopy.models.messaging import CategoryNew, StudentNew
from tutopy.services.exceptions import ValidationError


def _filled_template(service, path, students=(), categories=()):
    service.create_template(path)
    workbook = load_workbook(path)
    for row in students:
        workbook["Alumnes"].append(row)
    for row in categories:
        workbook["Categories"].append(row)
    workbook.save(path)


def _ods_file(path, students=(), categories=()):
    document = OpenDocumentSpreadsheet()
    for name, headers, rows in (
        ("Alumnes", ("Nom", "Cognoms", "Grup"), students),
        ("Categories", ("Nom",), categories),
    ):
        sheet = table.Table(name=name)
        for values in (headers, *rows):
            row = table.TableRow()
            for value in values:
                cell = table.TableCell(valuetype="string")
                cell.addElement(text.P(text=str(value)))
                row.addElement(cell)
            sheet.addElement(row)
        document.spreadsheet.addElement(sheet)
    document.save(str(path), addsuffix=False)


def test_template_te_els_fulls_i_capcaleres(db, tmp_path):
    service = create_services(db).bulk_import
    path = service.create_template(tmp_path / "plantilla")
    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["Instruccions", "Alumnes", "Categories"]
    assert tuple(cell.value for cell in workbook["Alumnes"][1]) == ("Nom", "Cognoms", "Grup")
    assert tuple(cell.value for cell in workbook["Categories"][1]) == ("Nom",)


def test_analyze_indica_full_fila_i_motiu(db, tmp_path):
    service = create_services(db).bulk_import
    path = tmp_path / "dades.xlsx"
    _filled_template(service, path, students=[("Anna", "Serra", "1r A"), (None, "Puig", "1r B")])
    preview = service.analyze(path)
    assert [str(issue) for issue in preview.issues] == [
        "Alumnes — fila 3: el nom és obligatori"
    ]


def test_importa_alumnes_i_reutilitza_categories_exactes(db, tmp_path):
    services = create_services(db)
    services.categories.create(CategoryNew("Acadèmic"))
    path = tmp_path / "dades.xlsx"
    _filled_template(services.bulk_import, path,
                     students=[("Anna", "Serra", "1r A")],
                     categories=[("ACADÈMIC",), ("Família",)])
    result = services.bulk_import.execute(services.bulk_import.analyze(path))
    assert result.students_created == 1
    assert result.categories_created == 1
    assert result.categories_reused == 1
    assert services.students.get_all()[0].group_name == "1r A"


def test_importa_alumnes_i_categories_des_d_ods(db, tmp_path):
    services = create_services(db)
    path = tmp_path / "dades.ods"
    _ods_file(path, students=[("Anna", "Serra", "1r A")],
              categories=[("Família",)])

    preview = services.bulk_import.analyze(path)
    result = services.bulk_import.execute(preview)

    assert preview.issues == ()
    assert result.students_created == 1
    assert result.categories_created == 1
    assert services.students.get_all()[0].filing_name == "Serra, Anna"


def test_importacio_rebutja_extensions_i_arxius_no_valids(db, tmp_path):
    service = create_services(db).bulk_import
    invalid = tmp_path / "dades.ods"
    invalid.write_text("no és un ODS")
    with pytest.raises(ValidationError, match="no és un fitxer vàlid"):
        service.analyze(invalid)
    unsupported = tmp_path / "dades.csv"
    unsupported.write_text("Nom,Cognoms")
    with pytest.raises(ValidationError, match="XLSX o ODS"):
        service.analyze(unsupported)


def test_coincidencia_similar_requereix_decisio_i_preserva_uuid(db, tmp_path):
    services = create_services(db)
    existing = services.students.create(StudentNew("Júlia", "Martínez", "2n A"))
    path = tmp_path / "dades.xlsx"
    _filled_template(services.bulk_import, path,
                     students=[("Julia", "Martines", "2n B")])
    preview = services.bulk_import.analyze(path)
    assert preview.conflicts[0].matches[0].id == existing.id
    result = services.bulk_import.execute(preview, (
        ImportDecision(2, ImportAction.UPDATE, existing.id),
    ))
    updated = services.students.get_by_id(existing.id)
    assert result.students_updated == 1
    assert updated.uuid == existing.uuid
    assert updated.group_name == "2n B"


def test_error_durant_execucio_desfa_tota_la_importacio(db, tmp_path):
    services = create_services(db)
    path = tmp_path / "dades.xlsx"
    _filled_template(services.bulk_import, path,
                     students=[("Anna", "Serra", "1r A"), ("Pau", "Puig", "1r B")])
    preview = services.bulk_import.analyze(path)
    original_create = services.students.create
    calls = 0

    def failing_create(data):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("error simulat")
        return original_create(data)

    services.students.create = failing_create
    try:
        # Els defectes inesperats no es converteixen en errors de domini, però
        # la transacció exterior continua garantint el rollback complet.
        with pytest.raises(RuntimeError, match="error simulat"):
            services.bulk_import.execute(preview)
    finally:
        services.students.create = original_create
    assert services.students.get_all() == []


def test_analisi_rebutja_fitxer_inexistent_i_massa_gran(db, tmp_path):
    service = create_services(db).bulk_import
    with pytest.raises(ValidationError, match="no existeix"):
        service.analyze(tmp_path / "absent.xlsx")

    path = tmp_path / "gran.xlsx"
    _filled_template(service, path)
    service.MAX_FILE_SIZE = 1
    with pytest.raises(ValidationError, match="20 MB"):
        service.analyze(path)


def test_analisi_detecta_fulls_capcaleres_formules_i_tipus_invalids(db, tmp_path):
    service = create_services(db).bulk_import
    path = tmp_path / "dades.xlsx"
    service.create_template(path)
    workbook = load_workbook(path)
    del workbook["Categories"]
    students = workbook["Alumnes"]
    students["A1"] = "Persona"
    students.append(["=A1", "Serra", True])
    workbook.save(path)

    preview = service.analyze(path)
    messages = [str(issue) for issue in preview.issues]

    assert any("capçaleres" in message for message in messages)
    assert any("falta el full" in message for message in messages)


def test_execucio_exigeix_decisions_i_valida_alumne_objectiu(db, tmp_path):
    services = create_services(db)
    existing = services.students.create(StudentNew("Júlia", "Martínez", "2n A"))
    path = tmp_path / "dades.xlsx"
    _filled_template(services.bulk_import, path,
                     students=[("Julia", "Martines", "2n B")])
    preview = services.bulk_import.analyze(path)

    with pytest.raises(ValidationError, match="Falten decisions"):
        services.bulk_import.execute(preview)
    with pytest.raises(ValidationError, match="coincidència vàlida"):
        services.bulk_import.execute(preview, (
            ImportDecision(2, ImportAction.UPDATE, existing.id + 100),
        ))


def test_execucio_permet_ometre_un_conflicte(db, tmp_path):
    services = create_services(db)
    services.students.create(StudentNew("Júlia", "Martínez", "2n A"))
    path = tmp_path / "dades.xlsx"
    _filled_template(services.bulk_import, path,
                     students=[("Julia", "Martines", "2n B")])
    preview = services.bulk_import.analyze(path)

    result = services.bulk_import.execute(preview, (
        ImportDecision(2, ImportAction.SKIP),
    ))

    assert result.students_skipped == 1
