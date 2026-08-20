from difflib import SequenceMatcher
import math
from pathlib import Path
import re
import unicodedata
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from odf import table as odf_table
from odf import teletype
from odf.opendocument import load as load_odf

from tutopy.models.bulk_import import (
    CategoryImportRow, ImportAction, ImportDecision, ImportIssue, ImportPreview,
    ImportResult, StudentConflict, StudentImportRow,
)
from tutopy.models.messaging import CategoryNew, Student, StudentNew
from tutopy.services.category_service import CategoryService
from tutopy.services.exceptions import DomainError, ValidationError
from tutopy.services.student_service import StudentService
from tutopy.services.validation_service import ValidationService


class BulkImportService:
    STUDENTS_SHEET = "Alumnes"
    CATEGORIES_SHEET = "Categories"
    STUDENT_HEADERS = ("Nom", "Cognoms", "Grup")
    CATEGORY_HEADERS = ("Nom",)
    MAX_FILE_SIZE = 20 * 1024 * 1024
    MAX_UNCOMPRESSED_SIZE = 100 * 1024 * 1024
    MAX_ARCHIVE_MEMBERS = 2_000
    MAX_ROWS_PER_SHEET = 10_000

    def __init__(self, students: StudentService, categories: CategoryService,
                 transaction_factory, similarity_threshold: float = 0.86):
        self.students = students
        self.categories = categories
        self.transaction_factory = transaction_factory
        self.similarity_threshold = similarity_threshold

    def create_template(self, destination: str | Path) -> Path:
        path = Path(destination)
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        workbook = Workbook()
        instructions = workbook.active
        instructions.title = "Instruccions"
        instructions.append(["Plantilla d’importació de Tutopy"])
        instructions.append(["Omple els fulls Alumnes i Categories. No canviïs les capçaleres."])
        instructions.append(["El nom i els cognoms són obligatoris; el grup és opcional."])
        students = workbook.create_sheet(self.STUDENTS_SHEET)
        students.append(self.STUDENT_HEADERS)
        categories = workbook.create_sheet(self.CATEGORIES_SHEET)
        categories.append(self.CATEGORY_HEADERS)
        for sheet in (students, categories):
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2B73B7")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        students.column_dimensions["A"].width = 22
        students.column_dimensions["B"].width = 32
        students.column_dimensions["C"].width = 18
        categories.column_dimensions["A"].width = 30
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path

    def analyze(self, source: str | Path) -> ImportPreview:
        path = Path(source)
        if not path.is_file():
            raise ValidationError("El full de càlcul seleccionat no existeix.")
        if path.suffix.lower() not in {".xlsx", ".ods"}:
            raise ValidationError("El format ha de ser XLSX o ODS.")
        self._validate_archive(path)
        try:
            workbook = (
                self._load_ods(path) if path.suffix.lower() == ".ods"
                else load_workbook(path, read_only=True, data_only=False)
            )
        # openpyxl i odfpy no exposen una jerarquia comuna d'errors de format.
        # Aquesta és una frontera d'entrada: qualsevol fallada del parser es
        # converteix en un error de domini sense continuar amb dades parcials.
        except Exception as error:
            raise ValidationError(
                "No s’ha pogut obrir el fitxer com a full de càlcul XLSX o ODS."
            ) from error

        issues: list[ImportIssue] = []
        student_rows: list[StudentImportRow] = []
        category_rows: list[CategoryImportRow] = []
        self._read_students(workbook, student_rows, issues)
        self._read_categories(workbook, category_rows, issues)
        existing = self.students.get_all()
        normalized_existing = tuple(
            (student, self._normalize(student.full_name)) for student in existing
        )
        name_index = self._name_index(normalized_existing)
        match_cache = {}
        conflicts = []
        for row in student_rows:
            incoming = self._normalize(row.full_name)
            matches = match_cache.get(incoming)
            if matches is None:
                matches = self._find_matches(incoming, name_index)
                match_cache[incoming] = matches
            if matches:
                conflicts.append(StudentConflict(row, matches))
        return ImportPreview(tuple(student_rows), tuple(category_rows),
                             tuple(conflicts), tuple(issues))

    def _validate_archive(self, path: Path) -> None:
        """Rebutja arxius massa grans o sospitosos abans de descomprimir-los."""
        try:
            if path.stat().st_size > self.MAX_FILE_SIZE:
                raise ValidationError("El full de càlcul supera el límit de 20 MB.")
            with ZipFile(path) as archive:
                members = archive.infolist()
                if len(members) > self.MAX_ARCHIVE_MEMBERS:
                    raise ValidationError("El full de càlcul conté massa fitxers interns.")
                if any(member.flag_bits & 0x1 for member in members):
                    raise ValidationError("No s’admeten fulls de càlcul xifrats.")
                if sum(member.file_size for member in members) > self.MAX_UNCOMPRESSED_SIZE:
                    raise ValidationError("El contingut descomprimit del full és massa gran.")
        except BadZipFile as error:
            raise ValidationError("El full de càlcul no és un fitxer vàlid.") from error

    def _load_ods(self, path: Path):
        document = load_odf(str(path))
        sheets = {}
        for sheet in document.spreadsheet.getElementsByType(odf_table.Table):
            name = sheet.getAttribute("name")
            sheets[name] = _OdsSheet(name, sheet, self.MAX_ROWS_PER_SHEET)
        return _OdsWorkbook(sheets)

    def execute(self, preview: ImportPreview,
                decisions: tuple[ImportDecision, ...] = ()) -> ImportResult:
        if preview.issues:
            raise ValidationError("\n".join(str(issue) for issue in preview.issues))
        decision_by_row = {decision.row: decision for decision in decisions}
        conflict_rows = {conflict.row.row: conflict for conflict in preview.conflicts}
        missing = sorted(set(conflict_rows) - set(decision_by_row))
        if missing:
            raise ValidationError(
                "Falten decisions per a les files d’alumnes: "
                + ", ".join(map(str, missing))
            )
        with self.transaction_factory():
            created, updated, skipped = self._execute_students(
                preview.students, decision_by_row, conflict_rows
            )
            categories_created, categories_reused = self._execute_categories(
                preview.categories
            )
        return ImportResult(created, updated, skipped, categories_created, categories_reused)

    def _execute_students(self, rows, decisions, conflicts):
        created = updated = skipped = 0
        for row in rows:
            decision = decisions.get(row.row, ImportDecision(row.row, ImportAction.CREATE))
            try:
                if decision.action == ImportAction.SKIP:
                    skipped += 1
                elif decision.action == ImportAction.UPDATE:
                    self._update_student(row, decision, conflicts.get(row.row))
                    updated += 1
                else:
                    self.students.create(StudentNew(row.name, row.surnames, row.group_name))
                    created += 1
            except (DomainError, ValueError, OSError) as error:
                raise ValidationError(
                    f"{self.STUDENTS_SHEET} — fila {row.row}: {error}"
                ) from error
        return created, updated, skipped

    def _update_student(self, row, decision, conflict) -> None:
        target = self.students.get_by_id(decision.student_id or 0)
        allowed_ids = {item.id for item in conflict.matches} if conflict else set()
        if target is None or target.id not in allowed_ids:
            raise ValidationError("l’alumne seleccionat no és una coincidència vàlida")
        self.students.update(Student(
            target.id, target.uuid, row.name, row.surnames, row.group_name
        ))

    def _execute_categories(self, rows):
        existing = {
            self._category_key(category.name): category
            for category in self.categories.get_all()
        }
        created = reused = 0
        for row in rows:
            key = self._category_key(row.name)
            if key in existing:
                reused += 1
                continue
            try:
                category = self.categories.create(CategoryNew(row.name))
                existing[key] = category
                created += 1
            except (DomainError, ValueError, OSError) as error:
                raise ValidationError(
                    f"{self.CATEGORIES_SHEET} — fila {row.row}: {error}"
                ) from error
        return created, reused

    def _read_students(self, workbook, output, issues) -> None:
        sheet = self._sheet(workbook, self.STUDENTS_SHEET, issues)
        if sheet is None or not self._headers(sheet, self.STUDENT_HEADERS, issues):
            return
        for number, values in enumerate(sheet.iter_rows(min_row=2, max_col=3, values_only=True), 2):
            if all(value is None or str(value).strip() == "" for value in values):
                continue
            cells = [self._cell(value, self.STUDENTS_SHEET, number, issues) for value in values]
            name, surnames, group = cells
            if not name:
                issues.append(ImportIssue(self.STUDENTS_SHEET, number, "el nom és obligatori"))
            if not surnames:
                issues.append(ImportIssue(self.STUDENTS_SHEET, number, "els cognoms són obligatoris"))
            if name and surnames:
                output.append(StudentImportRow(number, name, surnames, group))

    def _read_categories(self, workbook, output, issues) -> None:
        sheet = self._sheet(workbook, self.CATEGORIES_SHEET, issues)
        if sheet is None or not self._headers(sheet, self.CATEGORY_HEADERS, issues):
            return
        for number, (value,) in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), 2):
            if value is None or str(value).strip() == "":
                continue
            name = self._cell(value, self.CATEGORIES_SHEET, number, issues)
            if name:
                output.append(CategoryImportRow(number, name))

    def _sheet(self, workbook, name, issues):
        if name not in workbook.sheetnames:
            issues.append(ImportIssue(name, 1, f"falta el full «{name}»"))
            return None
        return workbook[name]

    def _headers(self, sheet, expected, issues) -> bool:
        actual = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1,
                                                                    max_col=len(expected))))
        if actual != expected:
            issues.append(ImportIssue(sheet.title, 1,
                                      "les capçaleres han de ser: " + ", ".join(expected)))
            return False
        return True

    @staticmethod
    def _cell(value, sheet, row, issues) -> str:
        if value is None:
            return ""
        if isinstance(value, str) and value.startswith("="):
            issues.append(ImportIssue(sheet, row, "no s’admeten fórmules"))
            return ""
        if not isinstance(value, (str, int, float)) or isinstance(value, bool):
            issues.append(ImportIssue(sheet, row, "el valor ha de ser text"))
            return ""
        return re.sub(r"\s+", " ", str(value)).strip()

    @staticmethod
    def _normalize(value: str) -> str:
        normalized = ValidationService.person_name(
            value, "El nom no pot estar buit."
        )
        decomposed = unicodedata.normalize("NFKD", normalized.casefold())
        return " ".join("".join(c for c in decomposed if not unicodedata.combining(c)).split())

    @staticmethod
    def _category_key(value: str) -> str:
        """Igualtat de categoria: majúscules i espais, però no accents ni grafies."""
        return " ".join(value.casefold().split())

    def _similar(self, row: StudentImportRow, student: Student) -> bool:
        incoming = self._normalize(row.full_name)
        current = self._normalize(student.full_name)
        return self._similar_names(incoming, current)

    def _similar_names(self, incoming: str, current: str) -> bool:
        if incoming == current:
            return True
        matcher = SequenceMatcher(None, incoming, current)
        return (
            matcher.quick_ratio() >= self.similarity_threshold
            and matcher.ratio() >= self.similarity_threshold
        )

    @staticmethod
    def _name_index(normalized_existing):
        """Agrupa noms per longitud conservant-ne l'ordre original."""
        by_length = {}
        for position, (student, name) in enumerate(normalized_existing):
            by_length.setdefault(len(name), []).append((position, student, name))
        return by_length

    def _find_matches(self, incoming: str, name_index) -> tuple[Student, ...]:
        """Busca només longituds que poden assolir el llindar de similitud."""
        threshold = self.similarity_threshold
        if threshold > 1:
            lengths = (len(incoming),)
        elif threshold <= 0:
            lengths = tuple(name_index)
        else:
            factor = (2 - threshold) / threshold
            minimum = math.ceil(len(incoming) / factor)
            maximum = math.floor(len(incoming) * factor)
            lengths = range(minimum, maximum + 1)
        candidates = [
            item
            for length in lengths
            for item in name_index.get(length, ())
        ]
        candidates.sort(key=lambda item: item[0])
        return tuple(
            student for _position, student, current in candidates
            if self._similar_names(incoming, current)
        )


class _CellValue:
    def __init__(self, value):
        self.value = value


class _OdsWorkbook:
    def __init__(self, sheets):
        self._sheets = sheets
        self.sheetnames = list(sheets)

    def __getitem__(self, name):
        return self._sheets[name]


class _OdsSheet:
    """Adaptador mínim a l'API de lectura que usa el servei d'importació."""

    def __init__(self, title, element, max_rows):
        self.title = title
        self._element = element
        self._max_rows = max_rows

    def iter_rows(self, min_row=1, max_row=None, max_col=None, values_only=False):
        current_row = 0
        for row in self._element.getElementsByType(odf_table.TableRow):
            repeated_rows = int(row.getAttribute("numberrowsrepeated") or 1)
            if repeated_rows > self._max_rows:
                raise ValidationError("El full conté massa files repetides.")
            values = self._row_values(row, max_col)
            for _ in range(repeated_rows):
                current_row += 1
                if current_row > self._max_rows:
                    raise ValidationError(
                        f"El full supera el límit de {self._max_rows} files."
                    )
                if current_row < min_row:
                    continue
                if max_row is not None and current_row > max_row:
                    return
                yield tuple(values if values_only else map(_CellValue, values))

    @staticmethod
    def _row_values(row, max_col):
        values = []
        cells = row.getElementsByType(odf_table.TableCell)
        for cell in cells:
            repeated = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            remaining = repeated if max_col is None else max(0, max_col - len(values))
            if remaining == 0:
                break
            formula = cell.getAttribute("formula")
            value = "=" + formula if formula else teletype.extractText(cell)
            values.extend([value or None] * min(repeated, remaining))
        if max_col is not None:
            values.extend([None] * (max_col - len(values)))
        return values
