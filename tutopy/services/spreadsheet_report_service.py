from collections import defaultdict
from datetime import date
from pathlib import Path
import unicodedata

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tutopy.database.daos.academic_course_dao import AcademicCourseDAO
from tutopy.database.daos.note_dao import NoteDAO
from tutopy.database.daos.student_dao import StudentDAO
from tutopy.database.daos.student_group_history_dao import StudentGroupHistoryDAO
from tutopy.services.exceptions import EntityNotFoundError, ValidationError
from tutopy.services.report_configuration_service import ReportConfigurationService
from tutopy.services.validation_service import ValidationService


class SpreadsheetReportService:
    """Genera l'informe XLSX de les notes de seguiment d'un alumne."""

    EXCEL_CELL_LIMIT = 32_767

    def __init__(self, students: StudentDAO, notes: NoteDAO,
                 courses: AcademicCourseDAO, group_history: StudentGroupHistoryDAO,
                 configuration: ReportConfigurationService, batch_loader=None):
        self.students = students
        self.notes = notes
        self.courses = courses
        self.group_history = group_history
        self.configuration = configuration
        self.batch_loader = batch_loader
        self.validation = ValidationService()

    def export_student(self, student_id: int, destination: str | Path,
                       include_terms: bool = False) -> Path:
        student_id = self.validation.positive_id(student_id)
        data = self.prepare_students([student_id], include_terms=include_terms)
        return self.export_prepared(student_id, destination, data, include_terms)

    def prepare_students(self, student_ids: list[int], include_terms: bool = False):
        """Carrega una vegada les dades compartides dels informes XLSX."""
        return self.batch_loader.load(
            student_ids, include_histories=True, include_terms=include_terms
        )

    def export_prepared(
        self, student_id: int, destination: str | Path, data,
        include_terms: bool = False,
    ) -> Path:
        """Genera un XLSX a partir d'un snapshot carregat prèviament."""
        student = data.students.get(student_id)
        if student is None:
            raise EntityNotFoundError("L’alumne seleccionat no existeix.")
        student_notes = sorted(
            data.notes[student_id], key=lambda note: (note.date, note.id)
        )
        if not student_notes:
            raise ValidationError("L’alumne no té notes per exportar.")
        path = Path(destination)
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        if not path.name:
            raise ValidationError("Cal indicar una destinació per a l’informe.")

        categories = data.categories
        histories = data.histories[student_id]
        by_course = defaultdict(list)
        for note in student_notes:
            by_course[note.course_id].append(note)
        course_names = self._course_names(by_course, data.course_names)

        workbook = Workbook()
        workbook.remove(workbook.active)
        for course_id, course_notes in sorted(
            by_course.items(), key=lambda item: course_names[item[0]]
        ):
            self._add_course_sheet(
                workbook, student, course_id, course_names[course_id],
                course_notes, histories,
                categories, include_terms, data.term_configurations,
            )

        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(path)
        except OSError as error:
            raise ValidationError("No s’ha pogut desar l’informe.") from error
        return path

    def _add_course_sheet(
        self, workbook, student, course_id, course_name, notes, histories,
        categories, include_terms: bool, term_configurations,
    ) -> None:
        sheet = workbook.create_sheet(self._sheet_title(course_name))
        rows = [(note, self._group_for(note.date, histories, student.group_name))
                for note in notes]
        groups = list(dict.fromkeys(group for _note, group in rows if group))
        last_column = len(categories) + (2 if include_terms else 1)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        title = sheet.cell(1, 1)
        self._set_text(title, f"{student.filing_name} — {', '.join(groups) or 'Sense grup'}")
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor="173A5E")
        title.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 25

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
        export_date = sheet.cell(2, 1)
        self._set_text(
            export_date, f"Data d’exportació: {date.today().strftime('%d/%m/%Y')}"
        )
        export_date.font = Font(italic=True, color="64748B")
        export_date.alignment = Alignment(horizontal="right")

        headers = (["Trimestre"] if include_terms else []) + ["Grup"] + [
            category.name for category in categories
        ]
        for column, header in enumerate(headers, 1):
            cell = sheet.cell(3, column)
            self._set_text(cell, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2B73B7")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        category_columns = {
            category.id: index
            for index, category in enumerate(categories, 3 if include_terms else 2)
        }
        for row_number, (note, group) in enumerate(rows, 4):
            group_column = self._write_term(
                sheet, row_number, course_id, group, note.date, term_configurations
            ) \
                if include_terms else 1
            self._set_text(sheet.cell(row_number, group_column), group)
            category_column = category_columns.get(note.category_id)
            if category_column is not None:
                display_date = date.fromisoformat(note.date).strftime("%d/%m/%Y")
                cell = sheet.cell(row_number, category_column)
                self._set_text(cell, f"{display_date} - {note.content}")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if include_terms:
            self._merge_consecutive_terms(sheet, 4, 3 + len(rows))
        self._format_sheet(sheet, last_column)

    def _write_term(
        self, sheet, row_number, course_id, group, note_date, term_configurations
    ) -> int:
        configuration = term_configurations.get((course_id, group)) if group else None
        if configuration is None:
            term = ""
        elif note_date < configuration.second_term_start:
            term = "1r"
        elif note_date < configuration.third_term_start:
            term = "2n"
        else:
            term = "3r"
        self._set_text(sheet.cell(row_number, 1), term)
        return 2

    @staticmethod
    def _course_names(by_course, names) -> dict[int, str]:
        missing = set(by_course) - names.keys()
        if missing:
            raise ValidationError("Una nota fa referència a un curs acadèmic inexistent.")
        return names

    @staticmethod
    def _group_for(note_date: str, histories, fallback: str) -> str:
        candidates = [
            history for history in histories
            if history.start_date <= note_date
            and (history.end_date is None or note_date <= history.end_date)
        ]
        if not candidates:
            return fallback
        return max(candidates, key=lambda history: (history.start_date, history.id)).group_name

    @staticmethod
    def _sheet_title(course_name: str) -> str:
        invalid = set('[]:*?/\\')
        title = "".join("-" if character in invalid else character for character in course_name)
        return title[:31] or "Curs"

    @classmethod
    def _safe_text(cls, value) -> str:
        """Prepara text Unicode vàlid per a una cel·la OOXML d'Excel."""
        text = unicodedata.normalize("NFC", "" if value is None else str(value))
        text = ILLEGAL_CHARACTERS_RE.sub("", text)
        return text[:cls.EXCEL_CELL_LIMIT]

    @classmethod
    def _set_text(cls, cell, value) -> None:
        """Força una cel·la a text, inclosos valors que comencen per ``=``."""
        cell.value = cls._safe_text(value)
        cell.data_type = "s"

    @staticmethod
    def _merge_consecutive_terms(sheet, first_row: int, last_row: int) -> None:
        start = first_row
        while start <= last_row:
            value = sheet.cell(start, 1).value
            end = start
            while end + 1 <= last_row and sheet.cell(end + 1, 1).value == value:
                end += 1
            if value and end > start:
                sheet.merge_cells(start_row=start, start_column=1,
                                  end_row=end, end_column=1)
                sheet.cell(start, 1).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            start = end + 1

    @staticmethod
    def _format_sheet(sheet, last_column: int) -> None:
        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A3:{get_column_letter(last_column)}{sheet.max_row}"
        for column in range(1, last_column + 1):
            letter = get_column_letter(column)
            sheet.column_dimensions[letter].width = 14 if column <= 2 else 34
        for row in sheet.iter_rows(min_row=4):
            for cell in row:
                if cell.alignment == Alignment():
                    cell.alignment = Alignment(vertical="top")
